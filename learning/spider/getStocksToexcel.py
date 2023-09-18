import easyquotation as eq
import openpyxl

file = '/home/frank/桌面/股票相关-2023.xlsx'
quotation = eq.use('sina')
qsxfj = quotation.real('600809')['600809']['close']
qrmw = quotation.real('603000')['603000']['close']
qgsyh = quotation.real('601398')['601398']['close']

tableAll = openpyxl.load_workbook(file)

tsxfj = tableAll['SXFJ-600809']
trmw = tableAll['RMW-603000']
tgsyh = tableAll['GSYH-601398']

print("山西汾酒："+str(qsxfj))
print("人民网："+str(qrmw))
print("工商银行："+str(qgsyh))

tsxfj.cell(5,7,qsxfj)
trmw.cell(5,7,qrmw)
tgsyh.cell(5,7,qgsyh)

tableAll.save(file)


