#coding=utf-8

"""
=============================================================
#   project: python
#      file: py01-requests-1.py
#    author: mianfeng.yang
#      date: 2019-09-27 17:30:58
=============================================================
"""
import re
import time

from openpyxl import Workbook
import requests
from scrapy import Selector


yd_hd = ['139', '138', '137', '136', '135', '134', '147', '150', '151', '152', '157', '158', '159', '172', '178', '182', '183', '184', '187', '188', '198']
yd_hd_int = [139, 138, 137, 136, 135, 134, 147, 150, 151, 152, 157, 158, 159, 172, 178, 182, 183, 184, 187, 188, 198]
yd_hd_139 = [1390015, 1390051, 1390158, 1390159, 1390514, 1390515, 1390516, 1390517, 1390518, 1390519, 1391290, 1391291, 1391292, 1391293, 1391294, 1391295, 1391296, 1391297, 1391298, 1391299, 1391300, 1391301, 1391302, 1391303, 1391304, 1391330, 1391331, 1391332, 1391333, 1391334, 1391335, 1391336, 1391337, 1391338, 1391339, 1391380, 1391381, 1391382, 1391383, 1391384, 1391385, 1391386, 1391387, 1391388, 1391389, 1391390, 1391391, 1391392, 1391393, 1391394, 1391395, 1391396, 1391397, 1391398, 1391399, 1391445, 1391446, 1391447, 1391448, 1391449, 1391470, 1391471, 1391472, 1391473, 1391474, 1391475, 1391476, 1391477, 1391478, 1391479, 1391590, 1391591, 1391592, 1391593, 1391594, 1391595, 1391596, 1391597, 1391598, 1391599, 1392140, 1392141, 1392142, 1392143, 1392144, 1395100, 1395101, 1395102, 1395103, 1395107, 1395108, 1395160, 1395161, 1395162, 1395163, 1395164, 1395165, 1395166, 1395167, 1395168, 1395169, 1395170, 1395171, 1395172, 1395173, 1395174, 1395175, 1395176, 1395177, 1395178, 1395179, 1395180, 1395181, 1395182, 1395183, 1395184, 1395185, 1395186, 1395187, 1395188, 1395189, 1395190, 1395191, 1395192, 1395193, 1395194, 1395195, 1395196, 1395197, 1395198, 1395199, 1395200, 1395201, 1395202, 1395203, 1395204, 1395205, 1395206, 1395207, 1395208, 1395209]
lt_hd = ['130', '131', '132', '140', '145', '146', '155', '156', '166', '167', '185', '186', '145', '175', '176']
lt_hd_int = [130, 131, 132, 140, 145, 146, 155, 156, 166, 167, 185, 186, 145, 175, 176]
dx_hd = ['133', '149', '153', '177', '173', '180', '181', '189', '191', '199']
dx_hd_int = [133, 149, 153, 177, 173, 180, 181, 189, 191, 199]

# dx_hd = [int(x) for x in dx_hd]
# req_text = requests.request(method='Get', url='https://www.jihaoba.com/haoduan/139/nanjing.htm').text
# sel = Selector(text=req_text)
# yd_hd_text = sel.xpath("//div[@class='hd_result']/div[1]/div[@class='hd_number']/a/text()").extract()
# lt_hd_text = sel.xpath("//div[@class='hd_result']/div[2]/div[@class='hd_number']/a/text()").extract()
# dx_hd_text = sel.xpath("//div[@class='hd_result']/div[3]/div[@class='hd_number']/a/text()").extract()

def Get_hd_3(hd_text, hd_list = []):
    for hd in hd_text:
        hd_3 = re.search("\d{3}", hd).group()
        hd_list.append(hd_3)
    return hd_list

# hd_text = sel.xpath("//ul[@class='hd-city']/li[@class='hd-city01']/a/text()").extract()
def Get_hd_7(req_text, hd_text):
    hd_text = [int(x) for x in hd_text]
    return hd_text

def Get_urls(hd_list, urls_list = []):
    for i in hd_list:
        url = 'https://www.jihaoba.com/haoduan/{}/nanjing.htm'.format(i)
        urls_list.append(url)
    return urls_list

def Parse_url(hd_list,res_list = []):
    for url in Get_urls(hd_list):
        req_text = requests.request(method='Get', url=url).text
        sel = Selector(text=req_text)
        hd_7 = sel.xpath("//ul[@class='hd-city']/li[@class='hd-city01']/a/text()").extract()
        for i in hd_7:
            res_list.append(i)
        #print(hd_7)
        time.sleep(2)
    return res_list


class WriteToExcel:
    """
    1. 在构造方法中，将生产者计算结果的队列传进来
    2. write_to_excel方法将队列中的数据写入excel文件
    """

    def __init__(self, res_list):
        # super().__init__()
        self.res_list = Parse_url(res_list)

    def write_to_excel(self):
        wb = Workbook()

        ws1 = wb.create_sheet("移动号段", 0)
        ws2 = wb.create_sheet("联通号段", 1)
        ws3 = wb.create_sheet("电信号段", 2)

        ws1['A1'] = "移动号段"
        ws2['A1'] = "联通号段"
        ws3['A1'] = "电信号段"


        # 不断从队列中取数据，直到队列取空不再取。一次取出的是一个列表，这样可以更好使用append方法批量向excel中写入，一次写一行
        if "1390015" in self.res_list:
            wb.active = 1
            ws1.append(self.res_list)
        if "1300006" in self.res_list:
            wb.active = 2
            ws2.append(self.res_list)
        if "1990158" in self.res_list:
            wb.active = 3
            ws3.append(self.res_list)


        # 保存文件
        wb.save('phone_hd.xlsx')
if __name__ == '__main__':
    # res_list = Parse_url(dx_hd)
    ws_res1 = WriteToExcel(dx_hd)
    ws_res1.write_to_excel()
    ws_res2 = WriteToExcel(lt_hd)
    ws_res2.write_to_excel()
    ws_res3 = WriteToExcel(yd_hd)
    ws_res3.write_to_excel()
    # Parse_url(yd_hd)
    # print(Parse_url(dx_hd))