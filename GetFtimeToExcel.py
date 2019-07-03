# coding=utf-8

"""
==============================================================
# @Time    : 2019/2/28 19:18
# @Author  : mifyang
# @Email   : mifyang@126.com
# @File    : GetFtimeToExcel
# @Software: PyCharm
==============================================================
"""
import os
import datetime
import threading
from openpyxl import Workbook
path1 = r'd:/soft/'
path2 = r'e:/'
path3 = r'd:/Users/frank/'
bpath1 = r'//192.168.1.11/pubin/'
bpath10 = r'//192.168.10.11/devin/'
depth = 4




def FindFiles(filepath, depth, child_list = []):
    '''
    递归遍历当前目录，返回当前目录下所有子目录列表
    参数 filepath为传入路径参数，child_list列表用于保存传入路径中每个子目录
    :param filepath:        传递一个路径
    :param depth            传递一个最大遍历目录的深度
    :param child_list:      用来保存子目录的列表
    :return: child_list     返回子目录列表
    '''
    depth -= 1
    child_list.append(filepath)
    for file in os.listdir(filepath):
        childpath = filepath + file + "/"
        # 如果超出最大深度，不再往下遍历
        if depth <= 0:
            continue
        # 如果有子目录则递归遍历子目录
        if os.path.isdir(childpath):
            FindFiles(childpath, depth)
    #返回所有子目录列表
    return child_list


def GetMaxDepth(dirlist, cdirdepth):
    '''
    找出当前目录子目录的最大深度并以字典的形式返回
    :param dirlist:     传递一个目录列表
    :param cdirdepth:   用于计算给定初始路径的 ”/“ 的个数
    :param dpath:       用于存放子目录的 ”/“ 的个数的列表
    :return: maxdpath   用于计算子目录的深度（"/"的个数）
    '''
    dpath_dict = {}
    dpath = []
    for i in dirlist:
        dpath.append(i.count("/"))
    maxdpath = max(dpath) - cdirdepth
    #将目录列表按”/“出现的个数进行排序，取出最后一个，也就是拥有最大目录深度的那个
    lpath = sorted(dirlist,key=lambda s: s.count("/"))[-1]
    #将目录的最大深度值，最大深度的目录存入字典
    dpath_dict[maxdpath] = lpath
    #由于函数是递归调用，所以再次调用时要清空字典，这样不会影响下次的值
    dirlist.clear()
    return dpath_dict


def GetMtime(dir):
    '''
    获取当前目录的修改时间并返回
    :param dir:         传递一个目录
    :return: filemtime  返回目录的修改时间
    '''
    filemtime = datetime.datetime.fromtimestamp(os.stat(dir).st_mtime).strftime('%Y-%m-%d %H:%M')
    year = datetime.datetime.fromtimestamp(os.stat(dir).st_mtime).year
    month = datetime.datetime.fromtimestamp(os.stat(dir).st_mtime).month
    print(year, month)

    return filemtime




def Pt(filepath, newestdir = {}):
    '''
    打印输出当目录及子目录和目录深度,返回目录及对应修改时间的字典,再把字典按时间排序取出最新一个
    :param filepath:        传递一个路径
    :param newestdir:       初始化一个空字典，用来保存最新修改的目录和对应的修改时间
    :return: newestdir
    '''

    #每次调用函数时清空字典
    newestdir.clear()
    dirlist = FindFiles(filepath, depth)
    #print("当前目录：{}".format(filepath))

    for cdir in dirlist:
        #print("\t包含子目录：{}\t最新修改时间：{}".format(cdir, GetMtime(cdir)))
        newestdir[cdir] = GetMtime(cdir)
    #result_dpath = GetMaxDepth(dirlist, filepath.count("/"))
    # for k, v in result_dpath.items():
    #     print("当前目录最大递归深度:{}\n最大深度目录：{}".format(k, v))
    #清空子目录列表，主要是不影响下次调用
    dirlist.clear()
    return newestdir


def GetFullPath_1(p):
    """
    根据传入的子目录名，生成完整路径
    :param p:       传入子目录列表中的一个元素
    :return:
    """
    rlist = []
    path = bpath1 + p + "/"
    rlist.append(path)
    if (p == "huoguangxin" or p == "霍广新"):
        rlist.append("霍广新")
    if (p == "hw.liu" or p == "刘宏伟"):
        rlist.append("刘宏伟")
    if (p == "jianhong.zhang" or p == "张建红"):
        rlist.append("张建红")
    if (p == "jili" or p == "吉利"):
        rlist.append("吉利")
    if (p == "libin" or p == "李宾"):
        rlist.append("李宾")
    if (p == "liujunwei" or p == "刘军伟"):
        rlist.append("刘军伟")
    if (p == "yuanjun" or p == "袁君"):
        rlist.append("袁君")
    if (p == "zhangyonghui" or p == "张永辉"):
        rlist.append("张永辉")
    if (p == "zhanglili" or p == "张丽丽"):
        rlist.append("张丽丽")
    if (p == "yangmianfeng" or p == "杨绵峰"):
        rlist.append("杨绵峰")
    if (p == "tangbaoyun" or p == "汤宝云"):
        rlist.append("汤宝云")
    result = sorted(Pt(path).items(), key=lambda d: d[1])[-1]
    #print("当前目录下最新修改的目录为是：{}\n".format(result))
    for i in result:
        rlist.append(i)
    #print(rlist)
    return rlist


def GetFullPath_10(p):
    """
       根据传入的子目录名，生成完整路径
       :param p:       传入子目录列表中的一个元素
       :return:
    """
    path = bpath10 + p + "/"
    rlist = []
    rlist.append(path)
    if (p == "huoguangxin" or p == "霍广新"):
        rlist.append("霍广新")
    if (p == "hw.liu" or p == "刘宏伟"):
        rlist.append("刘宏伟")
    if (p == "jianhong.zhang" or p == "张建红"):
        rlist.append("张建红")
    if (p == "jili" or p == "吉利"):
        rlist.append("吉利")
    if (p == "libin" or p == "李宾"):
        rlist.append("李宾")
    if (p == "liujunwei" or p == "刘军伟"):
        rlist.append("刘军伟")
    if (p == "yuanjun" or p == "袁君"):
        rlist.append("袁君")
    if (p == "zhangyonghui" or p == "张永辉"):
        rlist.append("张永辉")
    if (p == "zhanglili" or p == "张丽丽"):
        rlist.append("张丽丽")
    if (p == "yangmianfeng" or p == "杨绵峰"):
        rlist.append("杨绵峰")
    result = sorted(Pt(path).items(), key=lambda d: d[1])[-1]
    # print("当前目录下最新修改的目录为是：{}\n".format(result))
    for i in result:
        rlist.append(i)
    # print(rlist)
    return rlist

# def get_result_1(checklist1):
#     for i in checklist1:
#         t1 = threading.Thread(target=GetFullPath_1, args=(i, ))
#         t1.start()
#         t1.join()
#
# def get_result_10(checklist10):
#     for i in checklist10:
#         t10 = threading.Thread(target=GetFullPath_10, args=(i, ))
#         t10.start()
#         t10.join()


checklist10 = ['huoguangxin', '吉利', 'libin', 'liujunwei', 'yuanjun', '张建红', 'zhanglili','张永辉', 'hw.liu']
checklist1 = ["huoguangxin",  'jianhong.zhang', 'jili', 'libin',
              'liujunwei', 'yuanjun', 'zhangyonghui', '杨绵峰', 'zhanglili', 'hw.liu', '汤宝云']
checklistd = ['google', 'sublime3']

if __name__ == '__main__':
    start_time = datetime.datetime.now()
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "用户目录"
    ws['B1'] = "用户"
    ws['C1'] = "最新修改目录"
    ws['D1'] = "最新修改时间"
    ws['E1'] = "备份状态"
    ws['F1'] = "备注"
    for i in checklist1:
        tx1 = GetFullPath_1(i)
        ws.append(tx1)
    for i in checklist10:
        tx10 = GetFullPath_10(i)
        ws.append(tx10)
    
    wb.save(r'//192.168.1.11/pubin/杨绵峰/工作文件/备份检查/check_data_backup.xlsx')
    end_time = datetime.datetime.now()
    print("开始时间：{}\n结束时间：{}\n总共耗时：{}".format(start_time.strftime("%Y-%m-%d %H:%M:%S"),
                                             end_time.strftime("%Y-%m-%d %H:%M:%S"), end_time - start_time))