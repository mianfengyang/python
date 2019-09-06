
#coding:utf-8

import os
import datetime
from openpyxl import Workbook


bpath1 = r'//192.168.1.11/pubin/'
bpath10 = r'//192.168.10.11/devin/'
depth = 4

checklist10 = ['huoguangxin', '吉利', 'libin', 'liujunwei', 'yuanjun', '张建红', 'zhanglili', '张永辉', 'hw.liu']
checklist1 = ["huoguangxin", 'jianhong.zhang', 'jili', 'libin',
              'liujunwei', 'yuanjun', 'zhangyonghui', '杨绵峰', 'zhanglili', 'hw.liu', '汤宝云', '严建锋', '惠梦月']

def FindFiles(filepath, depth):
    '''
    递归遍历当前目录，返回当前目录下所有子目录列表
    参数 filepath为传入路径参数，child_list列表用于保存传入路径中每个子目录
    :param filepath:        传递一个路径
    :param depth            传递一个最大遍历目录的深度
    :param child_list:      用来保存子目录的列表
    :return: child_list     返回子目录列表
    '''
    depth -= 1
    yield filepath, os.stat(filepath).st_mtime
    for file in os.listdir(filepath):
        childpath = filepath + file + "/"
        # print(childpath)
        # 如果超出最大深度，不再往下遍历
        if depth <= 0:
            continue
        # 如果有子目录则递归遍历子目录
        if os.path.isdir(childpath):
            yield from FindFiles(childpath, depth)

def GetUser(u):
    """
    根据用户目录名返回中文用户名
    :return: user
    """
    user = ''
    if (u == "huoguangxin" or u == "霍广新"):
        user = "霍广新"
    if (u == "hw.liu" or u == "刘宏伟"):
        user = "刘宏伟"
    if (u == "jianhong.zhang" or u == "张建红"):
        user = "张建红"
    if (u == "jili" or u == "吉利"):
        user = "吉利"
    if (u == "libin" or u == "李宾"):
        user = "李宾"
    if (u == "liujunwei" or u == "刘军伟"):
        user = "刘军伟"
    if (u == "yuanjun" or u == "袁君"):
        user = "袁君"
    if (u == "zhangyonghui" or u == "张永辉"):
        user = "张永辉"
    if (u == "zhanglili" or u == "张丽丽"):
        user = "张丽丽"
    if (u == "yangmianfeng" or u == "杨绵峰"):
        user = "杨绵峰"
    if (u == "tangbaoyun" or u == "汤宝云"):
        user = "汤宝云"
    if (u == "huimengyue" or u == "惠梦月"):
        user = "惠梦月"
    if (u == "yanjianfeng" or u == "严建锋"):
        user = "严建锋"
    return user



def GetResult(bp, p):
    l = []
    res = FindFiles(bp + p + '/', depth)
    for i in res:
        l.append(i)
    # print(l)
    dd = sorted(l, key=lambda s: s[1])[-1]
    ff = (bp + p, dd[0], datetime.datetime.fromtimestamp(dd[1]).strftime('%Y-%m-%d %H:%M'), GetUser(p))
    return ff


def main():
    start_time = datetime.datetime.now()
    print("开始时间：{}".format(start_time.strftime("%Y-%m-%d %H:%M:%S")))
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "用户目录"
    ws['B1'] = "最新修改目录"
    ws['C1'] = "最新修改时间"
    ws['D1'] = "用户"
    ws['E1'] = "备份状态"
    ws['F1'] = "备注"

    for i in checklist1:
        tx1 = GetResult(bpath1, i)
        ws.append(tx1)

    for i in checklist10:
        tx10 = GetResult(bpath10, i)
        ws.append(tx10)

    # 根据表的行数，在指定列批量填入公式用于计算备份状态
    # 要在公式中使用变量，必须把公式打散再拼接
    # 此步在excel表中完成相对方便些

    cur_mon = datetime.datetime.now().month
    cur_yea = datetime.datetime.now().year
    for x in range(2, 24):
        ws.cell(row=x, column=5).value = "=IF(AND(YEAR(" + "C" + str(x) + ")=" + str(cur_yea) + ",MONTH(" + "C" + str(
            x) + ")>=" + str(cur_mon) + ")," + "\"已备份\"" + "," + "\"未备份\"" + ")"

    # 保存数据
    wb.save(r'D:/Desktop/杨绵峰/工作文件/备份检查/2019/x月检查情况.xlsx')

    end_time = datetime.datetime.now()
    print("结束时间：{}\n总共耗时：{}".format(end_time.strftime("%Y-%m-%d %H:%M:%S"), end_time - start_time))



if __name__ == "__main__":
    main()
