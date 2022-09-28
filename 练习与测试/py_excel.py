from openpyxl import Workbook
import datetime
wb = Workbook()
ws = wb.active
ws['A1'] = "用户目录"
ws['B1'] = "最新修改目录"
ws['C1'] = "最新修改时间"
ws['D1'] = "用户"
ws['E1'] = "备份状态"
ws['F1'] = "备注"

# print(datetime.datetime.now())

ws['c2'] = datetime.datetime.now()
# ws['E2'] = '=IF(AND(YEAR(C2)=2019,MONTH(C2)=7),"已备份","未备份")'
cur_mon = datetime.datetime.now().month
cur_yea = datetime.datetime.now().year
for x in range(2, 24):
    ws.cell(row=x, column=5).value = "=IF(AND(YEAR(" + "C" + str(x) + ")=" + str(cur_yea) + ",MONTH(" + "C" + str(
        x) + ")>=" + str(cur_mon) +")," + "\"已备份\"" + "," + "\"未备份\"" + ")"

wb.save(r'D:/Desktop/杨绵峰/工作文件/备份检查/2019/x月检查情况.xlsx')