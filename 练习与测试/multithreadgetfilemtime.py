
#coding:utf-8

import os
import datetime
from openpyxl import Workbook
import threading

bpath1 = r'//192.168.1.11/pubin/'
bpath10 = r'//192.168.10.11/devin/'
depth = 4

checklist10 = ['huoguangxin', '吉利', 'libin', 'liujunwei', 'yuanjun', '张建红', 'zhanglili', '张永辉', 'hw.liu']
checklist1 = ["huoguangxin", 'jianhong.zhang', 'jili', 'libin',
              'liujunwei', 'yuanjun', 'zhangyonghui', '杨绵峰', 'zhanglili', 'hw.liu', '汤宝云', '惠梦月', '严建锋']
testlist = ['google', 'Ghelper_1.4.6.beta', 'X220_Drivers', 'Xmanager Enterprise 5 Build 0987', 'xmind-8-update7-windows']

userdir = {}
userpaths = []


for i in checklist1:
    userpaths.append(bpath1 + i + '/')
for i in checklist10:
    userpaths.append(bpath10 + i + '/')




def find_child_dir(filepath, depth):
    '''
    递归遍历当前目录，返回当前目录下所有子目录列表
    参数 filepath为传入路径参数，child_list列表用于保存传入路径中每个子目录
    :param filepath:        传递一个路径
    :param depth            传递一个最大遍历目录的深度
    :yield: 利用生成器函数返回子目录列表
    '''

    depth -= 1
    yield filepath, os.stat(filepath).st_mtime
    for file in os.listdir(filepath):
        childpath = filepath + file + "/"
        # print(childpath)
        # 如果超出最大深度，不再往下遍历
        if depth <= 0:
            continue
        # 如果有子目录则递归遍历子目录
        if os.path.isdir(childpath):

            yield from find_child_dir(childpath, depth)

    # 返回所有子目录列表
    # return child_list





def GetMtime(dir):
    '''
    获取当前目录的修改时间并返回
    :param dir:         传递一个目录
    :return: filemtime  返回目录的修改时间
    '''
    filemtime = os.stat(dir).st_mtime
    return filemtime


def GetNewestDir(filepath):
    '''
    打印输出当目录及子目录和目录深度,返回目录及对应修改时间的字典,再把字典按时间排序取出最新一个
    :param filepath:        传递一个路径
    :param newestdir:       初始化一个空字典，用来保存最新修改的目录和对应的修改时间
    :return: newestdir
    '''
    # 每次调用函数时清空字典
    newestdir = []
    dirlist = FindFiles(filepath, depth)
    for i in range(len(dirlist) - 1):
        if GetMtime(dirlist[i]) > GetMtime(dirlist[i+1]):
            newestdir = [dirlist[i], GetMtime(dirlist[i])]
        else:
            newestdir = [dirlist[i+1], GetMtime(dirlist[i+1])]
    dirlist.clear()
    return newestdir


def GetUser(u):
    """
    根据用户目录名返回中文用户名
    :return: user
    """
    user = ''
    if (u == "huoguangxin" or u == "霍广新"):
        user = "霍广新"
    if (u == "hw.liu" or u == "刘宏伟"):
        user = "刘宏伟"
    if (u == "jianhong.zhang" or u == "张建红"):
        user = "张建红"
    if (u == "jili" or u == "吉利"):
        user = "吉利"
    if (u == "libin" or u == "李宾"):
        user = "李宾"
    if (u == "liujunwei" or u == "刘军伟"):
        user = "刘军伟"
    if (u == "yuanjun" or u == "袁君"):
        user = "袁君"
    if (u == "zhangyonghui" or u == "张永辉"):
        user = "张永辉"
    if (u == "zhanglili" or u == "张丽丽"):
        user = "张丽丽"
    if (u == "yangmianfeng" or u == "杨绵峰"):
        user = "杨绵峰"
    if (u == "tangbaoyun" or u == "汤宝云"):
        user = "汤宝云"
    if (u == "huimengyue" or u == "惠梦月"):
        user = "惠梦月"
    if (u == "yanjianfeng" or u == "严建锋"):
        user = "严建锋"
    return user



def GetResult(bp, p):
    userdir_dirtime = []
    res = find_child_dir(bp + p + '/', depth)
    for i in res:
        userdir_dirtime.append(i)
    # print(userdir_dirtime)
    newestdir = sorted(userdir_dirtime, key=lambda s: s[1])[-1]
    userdir_ndir_mtime_user = (bp + p, newestdir[0], datetime.datetime.fromtimestamp(newestdir[1]).strftime('%Y-%m-%d %H:%M'), GetUser(p))
    return userdir_ndir_mtime_user






if __name__ == "__main__":


    print("%s主线程开始……" % Process.name)


    start_time = datetime.datetime.now()
    print("开始时间：{}".format(start_time.strftime("%Y-%m-%d %H:%M:%S")))
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "用户目录"
    ws['B1'] = "最新修改目录"
    ws['C1'] = "最新修改时间"
    ws['D1'] = "用户"
    ws['E1'] = "备份状态"
    ws['F1'] = "备注"

    for i in checklist1:
        t1 = threading.Thread(target=GetResult, args=(bpath1, i))
        t1.start()

    for i in checklist10:
        t10 = threading.Thread(target=GetResult, args=(bpath10, i))
        t10.start()

    # 根据表的行数，在指定列批量填入公式用于计算备份状态
    # 要在公式中使用变量，必须把公式打散再拼接
    # 此步在excel表中完成相对方便些

    cur_mon = datetime.datetime.now().month
    cur_yea = datetime.datetime.now().year
    for x in range(2, 24):
        ws.cell(row=x, column=5).value = "=IF(AND(YEAR(" + "C" + str(x) + ")=" + str(cur_yea) + ",MONTH(" + "C" + str(
            x) + ")>=" + str(cur_mon) + ")," + "\"已备份\"" + "," + "\"未备份\"" + ")"

    # 保存数据
    wb.save(r'D:/Desktop/杨绵峰/工作文件/备份检查/2019/' + str(cur_mon) + '月检查情况.xlsx')


    end_time = datetime.datetime.now()
    print("结束时间：{}\n总共耗时：{}".format(end_time.strftime("%Y-%m-%d %H:%M:%S"), end_time - start_time))


