# coding=utf-8

"""
==============================================================
# @Time    : 2019/2/28 19:18
# @Author  : mifyang
# @Email   : mifyang@126.com
# @File    : GetFtimeToExcel
# @Software: PyCharm
==============================================================
"""
import os
import datetime
#下面这个库是个第三方库，需要安装，安装方法：cmd下 pip3 install openpyxl
from openpyxl import Workbook


bpath = r'//192.168.8.6/personal/'
depth = 4
user = ""

#定义要进行检查的用户文件夹
checklist = ["Yuhan.Mei", 'Yuliang.Liu', 'Yunfeng.Li', 'Zhaonan.Meng',
              'Zhe.Zhu', 'Zhen.Sun', 'Zhen.Han', 'Zhiyue.Deng', 'Zhongmin.Wang', 'Zhuo.Xie']

def FindFiles(filepath, depth, child_list=[]):
    '''
    递归遍历当前目录，返回当前目录下所有子目录列表
    参数 filepath为传入路径参数，child_list列表用于保存传入路径中每个子目录
    :param filepath:        传递一个路径
    :param depth            传递一个最大遍历目录的深度
    :param child_list:      用来保存子目录的列表
    :return: child_list     返回子目录列表
    '''
    depth -= 1
    child_list.append(filepath)
    for file in os.listdir(filepath):
        childpath = filepath + file + "/"
        # 如果超出最大深度，不再往下遍历
        if depth <= 0:
            continue
        # 如果有子目录则递归遍历子目录
        if os.path.isdir(childpath):
            FindFiles(childpath, depth)
    # 返回所有子目录列表
    return child_list



def GetMtime(dir):
    '''
    获取当前目录的修改时间并返回
    :param dir:         传递一个目录
    :return: filemtime  返回目录的修改时间
    '''
    filemtime = datetime.datetime.fromtimestamp(os.stat(dir).st_mtime).strftime('%Y-%m-%d %H:%M')
    return filemtime


def GetNewestDir(filepath, newestdir={}):
    '''
    打印输出当目录及子目录和目录深度,返回目录及对应修改时间的字典,再把字典按时间排序取出最新一个
    :param filepath:        传递一个路径
    :param newestdir:       初始化一个空字典，用来保存最新修改的目录和对应的修改时间
    :return: newestdir
    '''
    # 每次调用函数时清空字典
    newestdir.clear()
    dirlist = FindFiles(filepath, depth)
    for cdir in dirlist:
        newestdir[cdir] = GetMtime(cdir)
    result = sorted(newestdir.items(), key=lambda d: d[1])[-1]
    dirlist.clear()
    return result



def GetResult(p):
    """
    根据传入的子目录名，生成完整路径
    :param p:       传入子目录列表中的一个元素
    :return:
    """
    Ntlist = []
    path = bpath + p + "/"
    Ntlist.append(path)
    result = GetNewestDir(path)
    for i in result:
        Ntlist.append(i)
    return Ntlist


if __name__ == '__main__':
    start_time = datetime.datetime.now()
    print("开始时间：{}".format(start_time.strftime("%Y-%m-%d %H:%M:%S")))
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "用户目录"
    ws['B1'] = "最新修改目录"
    ws['C1'] = "最新修改时间"
    ws['D1'] = "备份状态"
    ws['E1'] = "备注"

    for i in checklist:
        tx1 = GetResult(i)
        ws.append(tx1)

    # 根据表的行数，在指定列批量填入公式用于计算备份状态
    # 要在公式中使用变量，必须把公式打散再拼接
    # 此步在excel表中完成相对方便些
    for x in range(2, 12):
        ws.cell(row=x, column=4).value = "=IF(AND(YEAR(" + "C" + str(x) + ")=2019,MONTH(" + "C" + str(
            x) + ")=7)," + "\"已备份\"" + "," + "\"未备份\"" + ")"
    #保存文件，写上路径和文件名
    wb.save(r'd:/check_data_backup.xlsx')


    end_time = datetime.datetime.now()
    print("结束时间：{}\n总共耗时：{}".format(end_time.strftime("%Y-%m-%d %H:%M:%S"), end_time - start_time))
