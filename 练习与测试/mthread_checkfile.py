# coding=utf-8

"""
==============================================================
# @Time    : 2019/9/20 19:52
# @Author  : mifyang
# @Email   : mifyang@126.com
# @File    : mthread_checkfile
# @Software: PyCharm
==============================================================
('c:/windows/appcompat/appraiser/AltData/', 1569149779.8340926)
('c:/windows/apppatch/zh-CN/', 1563119805.6274726)
('C:/Windows/Help/mui/', 1537027433.8218148)

"""
import os
from threading import Thread
from datetime import datetime
from queue import Queue
from openpyxl import Workbook

depth = 4
path = r'c:/windows/'
a = ['c:/windows/appcompat/', 'c:/windows/apppatch/', 'C:/Windows/Help/']
user_path_list = ['//192.168.1.11/pubin/huoguangxin/', '//192.168.1.11/pubin/jianhong.zhang/',
                  '//192.168.1.11/pubin/jili/', '//192.168.1.11/pubin/libin/', '//192.168.1.11/pubin/liujunwei/',
                  '//192.168.1.11/pubin/yuanjun/',
                  '//192.168.1.11/pubin/zhangyonghui/', '//192.168.1.11/pubin/杨绵峰/', '//192.168.1.11/pubin/zhanglili/',
                  '//192.168.1.11/pubin/hw.liu/', '//192.168.1.11/pubin/汤宝云/', '//192.168.1.11/pubin/严建锋/',
                  '//192.168.1.11/pubin/惠梦月/', '//192.168.10.11/devin/huoguangxin/', '//192.168.10.11/devin/吉利/',
                  '//192.168.10.11/devin/libin/', '//192.168.10.11/devin/liujunwei/', '//192.168.10.11/devin/yuanjun/',
                  '//192.168.10.11/devin/张建红/', '//192.168.10.11/devin/zhanglili/', '//192.168.10.11/devin/张永辉/',
                  '//192.168.10.11/devin/hw.liu/']


class Producer(Thread):
    """
    1. 构造方法中传入目录路径，遍历深度，队列
    2. 最终将计算结果放入队列
    """

    def __init__(self, filepath, depth, res_queue):
        super().__init__()
        self.filepath = filepath
        self.depth = depth
        self.res_queue = res_queue
        self.child_list = set()
        self.result = []
        self.cur_mon = datetime.now().month
        self.cur_yea = datetime.now().year

    def run(self):
        # 1. 获取用户根目录
        udir = self.get_user_path(self.filepath)
        self.result.append(udir)

        # 2. 找到用户目录里最新的mtime目录及mtime
        ndir_mtime = self.get_child_dir_mtime(self.filepath, self.depth)
        ndir, umtime = sorted(ndir_mtime, key=lambda s: s[1])[-1]
        self.result.append(ndir)
        str_umtime = datetime.fromtimestamp(umtime).strftime('%Y-%m-%d %H:%M')
        self.result.append(str_umtime)

        # 3. 返回用户目录的使用者
        user = self.get_user(self.filepath)
        self.result.append(user)

        # 4. 计算每用户的备份情况
        if (datetime.fromtimestamp(umtime).date().year == self.cur_yea) and (
                datetime.fromtimestamp(umtime).date().month >= self.cur_mon):
            self.result.append("已备份")
        else:
            self.result.append("未备份")

        # 5. 将完整列表放进队列
        self.res_queue.put(self.result)

    def get_user_path(self, filepath):
        """
        :param filepath:
        :return: 返回用户根目录路径
        """
        return filepath

    def get_child_dir_mtime(self, filepath, depth):
        """
        递归遍历用户目录及子目录，并将结果放入一个集合中，使用集合是因为其有自动去重功能
        :param filepath: 用户根目录
        :param depth: 遍历深度
        :return: 用户目录及深度范围内的所有路径和对应的最后修改时间戳
        """
        # 每递归一次深度减1
        depth -= 1
        # 向集合中添加元素，添加是一个元组，因为add方法一次只能添加1个元素
        self.child_list.add((filepath, os.stat(filepath).st_mtime))
        for file in os.listdir(filepath):
            childpath = filepath + file + "/"
            # 如果超出最大深度，不再往下遍历
            if depth <= 0:
                continue
            # 如果有子目录则递归遍历子目录child_list
            if os.path.isdir(childpath):
                self.get_child_dir_mtime(childpath, depth)
        # 返回所有子目录集合，利用了集合自动去重功能
        return self.child_list

    def get_user(self, filepath):
        """
        根据用户目录名返回中文用户名
        :return: user
        """
        user = ''
        if ("huoguangxin" in filepath) or ("霍广新" in filepath):
            user = "霍广新"
        if ("hw.liu" in filepath) or ("刘宏伟" in filepath):
            user = "刘宏伟"
        if ("jianhong.zhang" in filepath) or ("张建红" in filepath):
            user = "张建红"
        if ("jili" in filepath) or ("吉利" in filepath):
            user = "吉利"
        if ("libin" in filepath) or ("李宾" in filepath):
            user = "李宾"
        if ("liujunwei" in filepath) or ("刘军伟" in filepath):
            user = "刘军伟"
        if ("yuanjun" in filepath) or ("袁君" in filepath):
            user = "袁君"
        if ("zhangyonghui" in filepath) or ("张永辉" in filepath):
            user = "张永辉"
        if ("zhanglili" in filepath) or ("张丽丽" in filepath):
            user = "张丽丽"
        if ("杨绵峰" in filepath):
            user = "杨绵峰"
        if ("汤宝云" in filepath):
            user = "汤宝云"
        if ("惠梦月" in filepath):
            user = "惠梦月"
        if ("严建锋" in filepath):
            user = "严建锋"
        return user


class WriteToExcel:
    """
    1. 在构造方法中，将生产者计算结果的队列传进来
    2. write_to_excel方法将队列中的数据写入excel文件
    """

    def __init__(self, res_queue):
        # super().__init__()
        self.res_queue = res_queue

    def write_to_excel(self):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "用户目录"
        ws['B1'] = "最新修改目录"
        ws['C1'] = "最新修改时间"
        ws['D1'] = "用户"
        ws['E1'] = "备份状态"
        ws['F1'] = "备注"

        # 不断从队列中取数据，直到队列取空不再取。一次取出的是一个列表，这样可以更好使用append方法批量向excel中写入，一次写一行
        while True:
            ws.append(self.res_queue.get())
            if self.res_queue.empty():
                break

        # 保存文件
        wb.save(r'D:/Desktop/杨绵峰/工作文件/备份检查/2019/' + str(cur_mon) + '-backupcheck.xlsx')


if __name__ == '__main__':
    start_time = datetime.now()
    print("开始时间：{}".format(start_time.strftime("%Y-%m-%d %H:%M:%S")))

    # 创建一个队列
    res_queue = Queue()
    # 创建一个生产者线程列表，存放线程
    thread_list = []
    for i in userlist:
        t = Producer(i, depth, res_queue)
        thread_list.append(t)
    # 启动线程
    for i in thread_list:
        i.start()
    # 等待线程结束
    for j in thread_list:
        j.join()

    # 创建写入excel文件的对象
    w_res = WriteToExcel(res_queue)
    # 调用方法写入文件
    w_res.write_to_excel()

    # 获取当前月份
    cur_mon = datetime.now().month
    # 获取当前年份
    cur_yea = datetime.now().year

    # 利用excel的公式计算备份状态 这种方式非常的烦锁 直接使用python来解决比较好
    # for x in range(2, 24):
    #     ws.cell(row=x, column=5).value = "=IF(AND(YEAR(" + "C" + str(x) + ")=" + str(cur_yea) + ",MONTH(" + "C" + str(
    #         x) + ")>=" + str(cur_mon) + ")," + "\"已备份\"" + "," + "\"未备份\"" + ")"

    end_time = datetime.now()
    print("结束时间：{}\n总共耗时：{}".format(end_time.strftime("%Y-%m-%d %H:%M:%S"), end_time - start_time))
