# coding=utf-8

"""
==============================================================
# @Time    : 2019/2/28 19:18
# @Author  : mifyang
# @Email   : mifyang@126.com
# @File    : GetFtimeToExcel
# @Software: PyCharm
==============================================================
"""
import os
import datetime
from openpyxl import Workbook
import threading

bpath1 = r'//192.168.1.11/pubin/'
bpath10 = r'//192.168.10.11/devin/'
depth = 4
user = ""
checklist10 = ['huoguangxin', '吉利', 'libin', 'liujunwei', 'yuanjun', '张建红', 'zhanglili', '张永辉', 'hw.liu']
checklist1 = ["huoguangxin", 'jianhong.zhang', 'jili', 'libin',
              'liujunwei', 'yuanjun', 'zhangyonghui', '杨绵峰', 'zhanglili', 'hw.liu', '汤宝云', '严建锋', '惠梦月']

def FindFiles(filepath, depth, child_list=[]):
    '''
    递归遍历当前目录，返回当前目录下所有子目录列表
    参数 filepath为传入路径参数，child_list列表用于保存传入路径中每个子目录
    :param filepath:        传递一个路径
    :param depth            传递一个最大遍历目录的深度
    :param child_list:      用来保存子目录的列表
    :return: child_list     返回子目录列表
    '''
    depth -= 1
    child_list.append(filepath)
    for file in os.listdir(filepath):
        childpath = filepath + file + "/"
        # 如果超出最大深度，不再往下遍历
        if depth <= 0:
            continue
        # 如果有子目录则递归遍历子目录
        if os.path.isdir(childpath):
            FindFiles(childpath, depth)
    # 返回所有子目录列表
    return child_list


def GetMaxDepth(dirlist, cdirdepth):
    '''
    找出当前目录子目录的最大深度并以字典的形式返回
    :param diulist:     传递一个目录列表
    :param cdirdepth:   用于计算给定初始路径的 ”/“ 的个数
    :param dpath:       用于存放子目录的 ”/“ 的个数的列表
    :return: maxdpath   用于计算子目录的深度（"/"的个数）
    '''
    dpath_dict = {}
    dpath = []
    for i in dirlist:
        dpath.append(i.count("/"))
    maxdpath = max(dpath) - cdirdepth
    # 将目录列表按”/“出现的个数进行排序，取出最后一个，也就是拥有最大目录深度的那个
    lpath = sorted(dirlist, key=lambda s: s.count("/"))[-1]
    # 将目录的最大深度值，最大深度的目录存入字典
    dpath_dict[maxdpath] = lpath
    # 由于函数是递归调用，所以再次调用时要清空字典，这样不会影响下次的值
    dirlist.clear()
    return dpath_dict


def GetMtime(dir):
    '''
    获取当前目录的修改时间并返回
    :param dir:         传递一个目录
    :return: filemtime  返回目录的修改时间
    '''
    filemtime = datetime.datetime.fromtimestamp(os.stat(dir).st_mtime).strftime('%Y-%m-%d %H:%M')
    return filemtime


def GetNewestDir(filepath, newestdir={}):
    '''
    打印输出当目录及子目录和目录深度,返回目录及对应修改时间的字典,再把字典按时间排序取出最新一个
    :param filepath:        传递一个路径
    :param newestdir:       初始化一个空字典，用来保存最新修改的目录和对应的修改时间
    :return: newestdir
    '''
    # 每次调用函数时清空字典
    newestdir.clear()
    dirlist = FindFiles(filepath, depth)
    for cdir in dirlist:
        newestdir[cdir] = GetMtime(cdir)
    result = sorted(newestdir.items(), key=lambda d: d[1])[-1]
    dirlist.clear()
    return result

def GetUser(u):
    """
    根据用户目录名返回中文用户名
    :return: user
    """
    global user
    if (u == "huoguangxin" or u == "霍广新"):
        user = "霍广新"
    if (u == "hw.liu" or u == "刘宏伟"):
        user = "刘宏伟"
    if (u == "jianhong.zhang" or u == "张建红"):
        user = "张建红"
    if (u == "jili" or u == "吉利"):
        user = "吉利"
    if (u == "libin" or u == "李宾"):
        user = "李宾"
    if (u == "liujunwei" or u == "刘军伟"):
        user = "刘军伟"
    if (u == "yuanjun" or u == "袁君"):
        user = "袁君"
    if (u == "zhangyonghui" or u == "张永辉"):
        user = "张永辉"
    if (u == "zhanglili" or u == "张丽丽"):
        user = "张丽丽"
    if (u == "yangmianfeng" or u == "杨绵峰"):
        user = "杨绵峰"
    if (u == "tangbaoyun" or u == "汤宝云"):
        user = "汤宝云"
    if (u == "huimengyue" or u == "惠梦月"):
        user = "惠梦月"
    if (u == "yanjianfeng" or u == "严建锋"):
        user = "严建锋"
    return user


def GetResult_1(p):
    """
    根据传入的子目录名，生成完整路径
    :param p:       传入子目录列表中的一个元素
    :return:
    """
    Ntlist = []
    path = bpath1 + p + "/"
    Ntlist.append(path)
    result = GetNewestDir(path)
    for i in result:
        Ntlist.append(i)
    Ntlist.append(GetUser(p))
    # for i in Ntlist:
    #     ws.append(i)
    return Ntlist


def GetResult_10(p):
    """
       根据传入的子目录名，生成完整路径
       :param p:       传入子目录列表中的一个元素
       :return:
    """
    path = bpath10 + p + "/"
    ulist = []
    ulist.append(path)
    result = GetNewestDir(path)
    for i in result:
        ulist.append(i)
    ulist.append(GetUser(p))
    return ulist


if __name__ == '__main__':
    start_time = datetime.datetime.now()
    print("开始时间：{}".format(start_time.strftime("%Y-%m-%d %H:%M:%S")))
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "用户目录"
    ws['B1'] = "最新修改目录"
    ws['C1'] = "最新修改时间"
    ws['D1'] = "用户"
    ws['E1'] = "备份状态"
    ws['F1'] = "备注"
    for i in checklist1:
        tx1 = GetResult_1(i)
        ws.append(tx1)
    for i in checklist10:
        tx10 = GetResult_10(i)
        ws.append(tx10)

    for x in range(2, 24):
        ws.cell(row=x, column=5).value = "=IF(AND(YEAR(" + "C" + str(x) + ")=2019,MONTH(" + "C" + str(
        x) + ")=7)," + "\"已备份\"" + "," + "\"未备份\"" + ")"

    wb.save(r'//192.168.1.11/pubin/杨绵峰/工作文件/备份检查/check_data_backup.xlsx')

    end_time = datetime.datetime.now()
    print("结束时间：{}\n总共耗时：{}".format(end_time.strftime("%Y-%m-%d %H:%M:%S"), end_time - start_time))