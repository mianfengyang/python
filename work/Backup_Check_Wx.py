# coding=utf-8

"""
==============================================================
# @Time    : 2019/9/20 19:52
# @Author  : mifyang
# @Email   : mifyang@126.com
# @File    : mthread_checkfile
# @Software: PyCharm
==============================================================
"""
import os
from threading import Thread
from datetime import datetime
from queue import Queue
from openpyxl import Workbook

depth = 4
<<<<<<< HEAD
cur_mon = 11
=======
cur_mon = 12
>>>>>>> 879c9c01674cc147103065e98a58b947e57c5f47
user_path_list = ['//192.168.31.240/研发/王鼎/',
                  '//192.168.31.240/研发/聂莹霞/',
                  '//192.168.31.240/研发/钱峰/',
                  '//192.168.31.240/研发/王明慧/',
                  '//192.168.31.240/研发/郁苏阳/',
                  '//192.168.31.240/综合办/付军/',
                  '//192.168.31.240/综合办/颜蓉蓉/'
                  ]


class Producer(Thread):
	"""
	1. 构造方法中传入目录路径，遍历深度，队列
	2. 最终将计算结果放入队列
	"""
	
	def __init__(self, rootpath, depth, res_queue,cur_mon):
		super().__init__()
		self.rootpath = rootpath
		self.depth = depth
		self.res_queue = res_queue
		self.child_list = set()
		self.result = []
		self.cur_mon = cur_mon
		self.cur_yea = datetime.now().year
	
	def run(self):
		# 1. 获取用户根目录
		udir = self.get_user_path(self.rootpath)
		self.result.append(udir)
		
		# 2. 找到用户目录里最新的mtime目录及mtime
		ndir_mtime = self.get_child_dir_mtime(self.rootpath, self.depth)
		ndir, umtime = sorted(ndir_mtime, key=lambda s: s[1])[-1]
		self.result.append(ndir)
		str_umtime = datetime.fromtimestamp(umtime).strftime('%Y-%m-%d %H:%M')
		self.result.append(str_umtime)
		
		# 3. 返回用户目录的使用者
		#user = self.get_user(self.rootpath)
		#self.result.append(user)
		
		# 4. 计算每用户的备份情况
<<<<<<< HEAD
		if (datetime.fromtimestamp(umtime).date().year == self.cur_yea) and (
=======
		if (datetime.fromtimestamp(umtime).date().year >= self.cur_yea) and (
>>>>>>> 879c9c01674cc147103065e98a58b947e57c5f47
			datetime.fromtimestamp(umtime).date().month >= cur_mon):
			
			self.result.append("已备份")

		else:
			self.result.append("未备份")
		
		# 5. 将完整列表放进队列
		self.res_queue.put(self.result)
	
	def get_user_path(self, rootpath):
		"""
		:param rootpath:
		:return: 返回用户根目录路径
		"""
		return rootpath
	
	def get_child_dir_mtime(self, rootpath, depth):
		"""
		递归遍历用户目录及子目录，并将结果放入一个集合中，使用集合是因为其有自动去重功能
		:param rootpath: 用户根目录
		:param depth: 遍历深度
		:return: 用户目录及深度范围内的所有路径和对应的最后修改时间戳
		"""
		# 每递归一次深度减1
		depth -= 1
		# 向集合中添加元素，添加是一个元组，因为add方法一次只能添加1个元素
		self.child_list.add((rootpath, os.stat(rootpath).st_mtime))
		for file in os.listdir(rootpath):
			childpath = rootpath + file + "/"
			# 如果超出最大深度，不再往下遍历
			if depth <= 0:
				continue
			# 如果有子目录则递归遍历子目录child_list
			if os.path.isdir(childpath):
				self.get_child_dir_mtime(childpath, depth)
		# 返回所有子目录集合，利用了集合自动去重功能
		return self.child_list
	
	def get_user(self, rootpath):
		"""
		根据用户目录名返回中文用户名
		:return: user
		"""
		user = ''

		if ("王宇" in rootpath):
			user = "王宇"
		if ("聂莹霞" in rootpath):
			user = "聂莹霞"
		if ("王明慧" in rootpath):
			user = "王明慧"
		if ("郁苏阳" in rootpath):
			user = "郁苏阳"
		if ("吴静雨" in rootpath):
			user = "吴静雨"
		if ("王鼎" in rootpath):
			user = "王鼎"
		if ("钱峰" in rootpath):
			user = "钱峰"
		if ("付军" in rootpath):
			user = "付军"
		
		return user


class WriteToExcel:
	"""
	1. 在构造方法中，将生产者计算结果的队列传进来
	2. write_to_excel方法将队列中的数据写入excel文件
	"""
	
	def __init__(self, res_queue,cur_mon):
		# super().__init__()
		self.res_queue = res_queue
		self.cur_mon = cur_mon
		self.cur_yea = datetime.now().year
	
	def write_to_excel(self):
		wb = Workbook()
		ws = wb.active
		ws['A1'] = "用户目录"
		ws['B1'] = "最新修改目录"
		ws['C1'] = "最新修改时间"
		ws['D1'] = "备份状态"

		
		# 不断从队列中取数据，直到队列取空不再取。一次取出的是一个列表，这样可以更好使用append方法批量向excel中写入，一次写一行
		while True:
			ws.append(self.res_queue.get())
			if self.res_queue.empty():
				break
		
		# 保存文件
		wb.save(r'D:/Desktop/杨绵峰/工作文件/备份检查/无锡/2020/' + '无锡-' + str(cur_mon) + '月备份检查情况.xlsx')


if __name__ == '__main__':
	start_time = datetime.now()
	print("开始时间：{}".format(start_time.strftime("%Y-%m-%d %H:%M:%S")))
	
	# 创建一个队列
	res_queue = Queue()
	# 创建一个生产者线程列表，存放线程
	thread_list = []
	for i in user_path_list:
		t = Producer(i, depth, res_queue,cur_mon)
		thread_list.append(t)
	# 启动线程
	for i in thread_list:
		i.start()
	# 等待线程结束
	for j in thread_list:
		j.join()
	
	# 创建写入excel文件的对象
	w_res = WriteToExcel(res_queue,cur_mon)
	# 调用方法写入文件
	w_res.write_to_excel()
	
	
	end_time = datetime.now()
	print("结束时间：{}\n总共耗时：{}".format(end_time.strftime("%Y-%m-%d %H:%M:%S"), end_time - start_time))
